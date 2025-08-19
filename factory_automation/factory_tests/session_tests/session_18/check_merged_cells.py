#!/usr/bin/env python3
"""Check merged cells in Excel to understand the data structure"""

import pandas as pd
from openpyxl import load_workbook

def check_merged_cells():
    """Check merged cells in the Excel file"""
    
    file_path = '/Users/samarsingh/Factory_flow_Automation/inventory/ALLEN SOLLY (AS) STOCK 2026.xlsx'
    
    # Load workbook with openpyxl to check merged cells
    print("Loading workbook to check merged cells...")
    wb = load_workbook(file_path, data_only=True)
    
    # Check Sheet2
    if 'Sheet2' in wb.sheetnames:
        ws = wb['Sheet2']
        
        print("\nSheet2 Merged Cell Ranges:")
        for merged_range in ws.merged_cells.ranges:
            print(f"  {merged_range}")
            
        # Check specific rows around row 7 (Excel rows are 1-indexed, so row 7 is actually row 8 in Excel)
        print("\nChecking rows 60-75 in Sheet2 (around AS RELAXED CROP):")
        
        # Read with pandas but keep all rows
        df = pd.read_excel(file_path, sheet_name='Sheet2', header=0)
        
        # Look for rows 60-75
        print("\nData from rows 60-75:")
        subset = df.iloc[60:75] if len(df) > 75 else df.iloc[60:]
        
        for idx, row in subset.iterrows():
            trim_name = str(row.get('TRIM NAME', '')).strip()
            trim_code = str(row.get('TRIM CODE', '')).strip()
            size = str(row.get('SIZE', '')).strip()
            qty = str(row.get('QTY', '')).strip()
            
            # Print all rows, even if trim_name is NaN (merged cell)
            if trim_name == 'nan':
                print(f"Row {idx}: [MERGED] - {trim_code} - Size: {size}, Qty: {qty}")
            else:
                print(f"Row {idx}: {trim_name} - {trim_code} - Size: {size}, Qty: {qty}")
    
    # Try reading with different methods to handle merged cells
    print("\n" + "="*50)
    print("Reading with forward fill to handle merged cells:")
    
    df = pd.read_excel(file_path, sheet_name='Sheet2', header=0)
    
    # Forward fill the TRIM NAME column to handle merged cells
    df['TRIM NAME'] = df['TRIM NAME'].fillna(method='ffill')
    
    # Now find all AS RELAXED CROP WB entries
    relaxed_crop = df[df['TRIM NAME'].astype(str).str.contains('AS RELAXED CROP WB', case=False, na=False)]
    
    print("\nAS RELAXED CROP WB entries after forward fill:")
    if not relaxed_crop.empty:
        print(relaxed_crop[['TRIM NAME', 'TRIM CODE', 'SIZE', 'QTY']].to_string())
    else:
        print("No entries found")

if __name__ == "__main__":
    check_merged_cells()